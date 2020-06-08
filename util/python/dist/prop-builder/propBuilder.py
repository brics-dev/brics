#!/usr/bin/python3

import sys
import configparser
import argparse
import openpyxl
import collections
import os

import ConfigUtil
import CommonUtil
import simplejson as json

################################################
# createJobArgParser
################################################

def createJobArgParser():

    result = argparse.ArgumentParser()

    result.add_argument('-p', '--props', help='Properties file path', required=False)
    result.add_argument('-m', '--mode', help='Execution mode (optional)', required=False, default="manual")
    #result.add_argument('-r', '--reset', help='Reset', action="store_true")

    result.add_argument('-e', '--env', help='Environment', required=True)

    result.add_argument('-l', '--logDir', help='Directory to output log files', required=False)
    result.add_argument('-i', '--instFile', help='Instructions file', required=False)
    result.add_argument('-d', '--workDir', help='Work dir', required=False)
    result.add_argument('-w', '--passFile', help='Password file', required=False)

    return result

################################################
# getSheetEnvironmentLookup
################################################

def getSheetEnvironmentLookup(argSheet):

    envColNo = 2

    result = {}

    while True:
        envId = argSheet.cell(row=1, column=envColNo).value
        if envId is None: break
        #print("envId: <%s>" % envId)
        result[envId] = envColNo
        envColNo = envColNo + 1

    return result

################################################
# getFileProperties
################################################

def getFileProperties(argSheet, argEnvCol):

    envRowNo = 2

    #using OrderedDict to preserve the ordering of propery listing from the matrix file
    result = collections.OrderedDict()

    while True:
        propName = argSheet.cell(row=envRowNo, column=1).value
        if propName is None: break

        if argEnvCol is not None:
            locVal = sheet.cell(row=envRowNo, column=argEnvCol).value
            result[propName] = locVal

        envRowNo = envRowNo + 1

    return result

################################################
# main
################################################

if __name__ == '__main__':

    argparser = createJobArgParser()

    args = argparser.parse_args()

    propFile = args.props
    execMode = args.mode
    env = args.env

    if not propFile is None:
        config = configparser.ConfigParser()
        config.read(propFile)

        fileSysConfHash = config['FILE-SYS']

    else:

        fileSysConfHash = {}
        fileSysConfHash['logDir'] = args.logDir
        fileSysConfHash['instFile'] = args.instFile
        fileSysConfHash['workDir'] = args.workDir
        fileSysConfHash['passFile'] = args.passFile

    fileSysConf = ConfigUtil.getFileSysConfig(fileSysConfHash)

    matrixFile = fileSysConf.instFile

    discard, matrixFileExt = os.path.splitext(matrixFile)

    print("matrixFileExt: <%s>" % matrixFileExt)

    if matrixFileExt == ".xlsx":
        print("Matrix file: <%s>" % matrixFile)

        workbook = openpyxl.load_workbook(matrixFile)

        print("Matrix file opened")

        sheetNames = workbook.sheetnames

        for sheetName in sheetNames:

            print("*** Sheet: <%s>" % sheetName)

            sheet = workbook[sheetName]

            sheetEnvLookup = getSheetEnvironmentLookup(sheet)

            sheetEnvCol = sheetEnvLookup.get(env)

            fileProps = getFileProperties(sheet, sheetEnvCol)

            for prop, val in fileProps.items():
                print("%s=%s" % (prop, val))

    elif matrixFileExt == ".json":

        propJson = CommonUtil.getFileContentString(matrixFile)
        parsed_json = json.loads(propJson)
        matrix = parsed_json.get("matrix").get("modules.properties")

        for prop, val in matrix.items():
            print("%s=%s" % (prop, val))
